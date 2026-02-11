from fastapi import APIRouter, Query, Request, Form, Depends, HTTPException, status
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import and_
from typing import Optional
from datetime import date, datetime
import bcrypt
from starlette.status import HTTP_303_SEE_OTHER
from app.config.database import get_db
from app.schema import models, schemas
from app.repository import crud


from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO



router = APIRouter()
templates = Jinja2Templates(directory="app/templates")

# ============================================================================
# FUNCIONES DE AUTENTICACIÓN
# ============================================================================

def hashear_password(password: str) -> str:
    """Hashea una contraseña usando bcrypt"""
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verificar_password(plain_password: str, hashed_password: str) -> bool:
    """Verifica si una contraseña coincide con su hash"""
    return bcrypt.checkpw(plain_password.encode('utf-8'), hashed_password.encode('utf-8'))

# ============================================================================
# RUTAS DE AUTENTICACIÓN
# ============================================================================

@router.get("/", response_class=HTMLResponse)
def root(request: Request):
    """Página principal (login)"""
    return templates.TemplateResponse("login.html", {"request": request})

@router.get("/login", response_class=HTMLResponse)
def mostrar_login(request: Request):
    """Mostrar formulario de login"""
    return templates.TemplateResponse("login.html", {"request": request})

@router.post("/login")
async def login(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db)
):
    """Procesar login de usuario"""
    usuario = crud.obtener_usuario_por_username(db, username)
    
    if usuario and verificar_password(password, usuario.password):
        request.session["usuario_id"] = usuario.id
        return RedirectResponse(url="/dashboard", status_code=HTTP_303_SEE_OTHER)

    return templates.TemplateResponse("login.html", {
        "request": request,
        "error": "Credenciales inválidas"
    })

@router.get("/logout")
def logout(request: Request):
    """Cerrar sesión del usuario"""
    request.session.clear()
    return RedirectResponse(url="/", status_code=303)


# ============================================================================
# RUTAS DE REGISTRO
# ============================================================================

# ============================================================================
# RUTAS DE REGISTRO
# ============================================================================

# ============================================================================
# RUTAS DE REGISTRO
# ============================================================================

@router.get("/register", response_class=HTMLResponse)
def mostrar_registro(request: Request):
    """Mostrar formulario de registro"""
    return templates.TemplateResponse("register.html", {"request": request})

@router.post("/register")
async def registrar_usuario(
    request: Request,
    nombre: str = Form(...),
    email: str = Form(...),
    username: str = Form(...),
    password: str = Form(...),
    confirm_password: str = Form(...),  # Cambia el nombre para que coincida con el formulario
    db: Session = Depends(get_db)
):
    """Procesar registro de nuevo usuario"""
    
    try:
        print(f"\n📝 INTENTANDO REGISTRAR USUARIO:")
        print(f"  Nombre: {nombre}")
        print(f"  Email: {email}")
        print(f"  Username: {username}")
        
        # 1. Verificar que las contraseñas coincidan
        if password != confirm_password:
            print("❌ Las contraseñas no coinciden")
            return templates.TemplateResponse("register.html", {
                "request": request,
                "error": "Las contraseñas no coinciden"
            })
        
        # 2. Validar longitud mínima de contraseña
        if len(password) < 6:
            print(f"❌ Contraseña muy corta: {len(password)} caracteres")
            return templates.TemplateResponse("register.html", {
                "request": request,
                "error": "La contraseña debe tener al menos 6 caracteres"
            })
        
        # 3. Verificar si el usuario ya existe
        usuario_existente = crud.obtener_usuario_por_username(db, username)
        if usuario_existente:
            print(f"❌ Usuario '{username}' ya existe")
            return templates.TemplateResponse("register.html", {
                "request": request,
                "error": "El nombre de usuario ya está en uso"
            })
        
        # 4. Verificar si el email ya existe
        email_existente = crud.obtener_usuario_por_email(db, email)
        if email_existente:
            print(f"❌ Email '{email}' ya registrado")
            return templates.TemplateResponse("register.html", {
                "request": request,
                "error": "El email ya está registrado"
            })
        
        # 5. Crear el usuario
        print("✅ Validaciones pasadas, creando usuario...")
        usuario_data = schemas.UsuarioCreate(
            nombre=nombre,
            email=email,
            username=username,
            password=password
        )
        
        nuevo_usuario = crud.crear_usuario(db, usuario_data)
        print(f"✅ Usuario creado exitosamente con ID: {nuevo_usuario.id}")
        
        # 6. Mostrar mensaje de éxito y redirigir al login después de 3 segundos
        return templates.TemplateResponse("register.html", {
            "request": request,
            "success": "¡Registro exitoso! Serás redirigido al login en 8 segundos...",
            "redirect": True
        })
        
    except Exception as e:
        print(f"\n❌ ERROR CRÍTICO AL REGISTRAR USUARIO: {str(e)}")
        import traceback
        traceback.print_exc()
        return templates.TemplateResponse("register.html", {
            "request": request,
            "error": f"Error al crear la cuenta: {str(e)}"
        })

# ============================================================================
# DASHBOARD
# ============================================================================

@router.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request, db: Session = Depends(get_db)):
    """Mostrar dashboard principal"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=303)
    
    usuario = crud.obtener_usuario_por_id(db, usuario_id)
    if not usuario:
        return RedirectResponse(url="/login", status_code=303)

    stats = crud.obtener_estadisticas_dashboard(db, usuario_id)
    
    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "usuario": usuario,
        "stats": stats,
        "fecha_actual": datetime.now()
    })

# ============================================================================
# RUTAS DE INGRESOS
# ============================================================================

@router.get("/ingresos")
def listar_ingresos(
    request: Request, 
    db: Session = Depends(get_db), 
    page: int = 1,
    tipo: Optional[str] = None,
    estado: Optional[str] = None  # Cambia 'pagado' por 'estado'
):
    """
    Listar ingresos con paginación y filtros
    
    Args:
        request: Request de FastAPI
        db: Sesión de base de datos
        page: Número de página (por defecto 1)
        tipo: Tipo de categoría para filtrar (fijo, variable, opcional)
        estado: Estado para filtrar ('recibido', 'pendiente', o None)
    """
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=303)
    
    print(f"\n=== LISTANDO INGRESOS PARA USUARIO ID: {usuario_id} ===")
    print(f"Filtros recibidos - tipo: '{tipo}', estado: '{estado}'")
    
    # Validar que el estado sea 'recibido' o 'pendiente'
    estado_filtro = None
    if estado is not None and estado != "":
        if estado.lower() == 'recibido':
            estado_filtro = 'recibido'
        elif estado.lower() == 'pendiente':
            estado_filtro = 'pendiente'
        else:
            print(f"⚠️  Valor inválido para 'estado': '{estado}', ignorando filtro")
    
    print(f"Filtro 'estado' procesado: {estado_filtro}")
    
    # Obtener ingresos del usuario actual
    data = crud.obtener_ingresos_paginados(
        db, 
        usuario_id,
        page=page, 
        tipo=tipo,
        estado=estado_filtro  # Cambia 'pagado' por 'estado'
    )
    
    print(f"Total ingresos encontrados: {len(data['ingresos'])}")
    
    mensaje = request.session.pop("mensaje", None)
    
    return templates.TemplateResponse(
        "ingresos_listado.html",
        {
            "request": request,
            "ingresos": data["ingresos"],
            "total_pages": data["total_pages"],
            "current_page": page,
            "filtro_tipo": tipo,
            "filtro_estado": estado,  # Mantener el string original para el template
            "mensaje": mensaje
        }
    )


@router.get("/ingresos/nuevo", response_class=HTMLResponse)
def formulario_nuevo_ingreso(request: Request, db: Session = Depends(get_db)):
    """Mostrar formulario para crear nuevo ingreso"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=303)
    
    # Obtener categorías existentes para autocompletar
    categorias_existentes = crud.obtener_categorias(db, usuario_id)
    
    # Fecha actual para el formulario
    hoy = date.today().isoformat()
    
    return templates.TemplateResponse("ingreso_form.html", {
        "request": request,
        "ingreso": None,  # Indica que es un nuevo ingreso
        "categorias_existentes": categorias_existentes,
        "hoy": hoy
    })

@router.post("/ingresos/crear")
def crear_ingreso(
    request: Request,
    valor: float = Form(...),
    fecha: str = Form(...),
    categoria: str = Form(...),
    tipo: str = Form("variable"),
    estado: str = Form("pendiente"),
    notas: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    """Crear un nuevo ingreso"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=303)

    try:
        print(f"\n{'='*60}")
        print(f"➕ CREANDO NUEVO INGRESO")
        print(f"{'='*60}")
        
        # Validaciones básicas
        if not fecha or fecha.strip() == "":
            raise ValueError("❌ ERROR: La fecha es requerida")
        
        fecha_dt = date.fromisoformat(fecha)
        categoria_nombre = categoria.strip()
        
        if not categoria_nombre:
            raise ValueError("❌ ERROR: El nombre de categoría es requerido")
        
        if tipo not in ['fijo', 'variable', 'opcional']:
            tipo = 'variable'
        
        if estado not in ['pendiente', 'recibido']:
            estado = 'pendiente'
        
        # Buscar o crear categoría
        categoria_existente = db.query(models.Categoria).filter(
            models.Categoria.nombre == categoria_nombre,
            models.Categoria.usuario_id == usuario_id
        ).first()
        
        if categoria_existente:
            print(f"✅ Usando categoría existente: '{categoria_existente.nombre}'")
            categoria_id = categoria_existente.id
        else:
            print(f"➕ Creando nueva categoría: '{categoria_nombre}'")
            nueva_categoria = models.Categoria(
                nombre=categoria_nombre,
                tipo=tipo,
                usuario_id=usuario_id
            )
            db.add(nueva_categoria)
            db.flush()
            categoria_id = nueva_categoria.id
        
        # Crear el ingreso
        nuevo_ingreso = models.Ingreso(
            valor=valor,
            fecha=fecha_dt,
            categoria_id=categoria_id,
            estado=estado,
            notas=notas,
            usuario_id=usuario_id,
            es_salario=False,
            recurrente=False
        )
        
        db.add(nuevo_ingreso)
        db.commit()
        
        print(f"✅ Ingreso creado exitosamente con ID: {nuevo_ingreso.id}")
        
        request.session['mensaje'] = {
            'tipo': 'exito',
            'titulo': '¡Éxito!',
            'texto': 'Ingreso creado exitosamente'
        }
        
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        db.rollback()
        request.session['mensaje'] = {
            'tipo': 'error',
            'titulo': 'Error',
            'texto': f'Error: {str(e)}'
        }

    return RedirectResponse(url="/ingresos", status_code=303)

@router.get("/ingresos/editar/{ingreso_id}", response_class=HTMLResponse)
def formulario_editar_ingreso(
    request: Request,
    ingreso_id: int,
    db: Session = Depends(get_db)
):
    """Mostrar formulario para editar un ingreso existente"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=303)

    # Obtener el ingreso a editar
    ingreso = crud.obtener_ingreso(db, ingreso_id)
    if not ingreso or ingreso.usuario_id != usuario_id:
        raise HTTPException(status_code=404, detail="Ingreso no encontrado")
    
    # Obtener categorías existentes
    categorias_existentes = crud.obtener_categorias(db, usuario_id)
    
    # Fecha actual (para consistencia en el template)
    hoy = date.today().isoformat()

    return templates.TemplateResponse("ingreso_form.html", {
        "request": request,
        "ingreso": ingreso,
        "categorias_existentes": categorias_existentes,
        "hoy": hoy
    })

@router.post("/ingresos/editar")
def editar_ingreso(
    request: Request,
    id: int = Form(...),
    valor: float = Form(...),
    fecha: str = Form(...),
    categoria: str = Form(...),
    tipo: str = Form("variable"),
    estado: str = Form("pendiente"),
    notas: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=303)

    try:
        print(f"\n{'='*60}")
        print(f"📝 INICIANDO EDICIÓN DE INGRESO ID: {id}")
        print(f"{'='*60}")
        
        # ============================================
        # 1. VALIDACIÓN DE DATOS DE ENTRADA
        # ============================================
        print(f"\n📋 DATOS RECIBIDOS DEL FORMULARIO:")
        print(f"  - ID: {id}")
        print(f"  - Valor: {valor}")
        print(f"  - Fecha: {fecha}")
        print(f"  - Categoría: '{categoria}'")
        print(f"  - Tipo: '{tipo}'")
        print(f"  - Estado: '{estado}'")
        print(f"  - Notas: '{notas}'")
        
        # Validaciones...
        if not fecha or fecha.strip() == "":
            raise ValueError("❌ ERROR: La fecha es requerida")
        
        fecha_dt = date.fromisoformat(fecha)
        categoria_nombre = categoria.strip()
        
        if not categoria_nombre:
            raise ValueError("❌ ERROR: El nombre de categoría es requerido")
        
        if tipo not in ['fijo', 'variable', 'opcional']:
            tipo = 'variable'
        
        if estado not in ['pendiente', 'recibido']:
            estado = 'pendiente'
        
        # ============================================
        # 2. OBTENER EL INGRESO EXISTENTE
        # ============================================
        ingreso = db.query(models.Ingreso).filter(
            models.Ingreso.id == id,
            models.Ingreso.usuario_id == usuario_id
        ).first()
        
        if not ingreso:
            raise HTTPException(status_code=404, detail="Ingreso no encontrado")
        
        print(f"\n📄 INGRESO ACTUAL EN BD:")
        print(f"  - ID: {ingreso.id}")
        print(f"  - Categoría ID actual: {ingreso.categoria_id}")
        
        # Obtener categoría actual
        categoria_actual = db.query(models.Categoria).filter(
            models.Categoria.id == ingreso.categoria_id
        ).first() if ingreso.categoria_id else None
        
        if categoria_actual:
            print(f"  - Categoría actual: '{categoria_actual.nombre}' ({categoria_actual.tipo})")
        
        # ============================================
        # 3. MANEJO DE CATEGORÍA (CORREGIDO - NO MODIFICA EXISTENTES)
        # ============================================
        print(f"\n🔍 MANEJO DE CATEGORÍA:")
        print(f"   Nombre solicitado: '{categoria_nombre}'")
        print(f"   Tipo solicitado: '{tipo}'")
        
        # IMPORTANTE: Siempre buscar categoría por NOMBRE + TIPO
        # NO modificamos categorías existentes, buscamos o creamos una nueva
        
        # Buscar si ya existe una categoría con este nombre Y tipo
        categoria_existente = db.query(models.Categoria).filter(
            models.Categoria.nombre == categoria_nombre,
            models.Categoria.tipo == tipo,
            models.Categoria.usuario_id == usuario_id
        ).first()
        
        if categoria_existente:
            print(f"✅ Encontrada categoría existente:")
            print(f"   ID: {categoria_existente.id}")
            print(f"   Nombre: '{categoria_existente.nombre}'")
            print(f"   Tipo: '{categoria_existente.tipo}'")
            categoria_id = categoria_existente.id
        else:
            # Crear NUEVA categoría con nombre + tipo
            print(f"➕ Creando NUEVA categoría:")
            print(f"   Nombre: '{categoria_nombre}'")
            print(f"   Tipo: '{tipo}'")
            
            # Verificar si ya existe una categoría con este nombre pero diferente tipo
            misma_nombre_diferente_tipo = db.query(models.Categoria).filter(
                models.Categoria.nombre == categoria_nombre,
                models.Categoria.usuario_id == usuario_id,
                models.Categoria.tipo != tipo
            ).first()
            
            if misma_nombre_diferente_tipo:
                print(f"⚠️  Ya existe categoría '{categoria_nombre}' con tipo '{misma_nombre_diferente_tipo.tipo}'")
                print(f"   Creando nueva categoría con tipo diferente")
            
            nueva_categoria = models.Categoria(
                nombre=categoria_nombre,
                tipo=tipo,
                usuario_id=usuario_id
            )
            db.add(nueva_categoria)
            db.flush()
            categoria_id = nueva_categoria.id
            
            print(f"✅ Nueva categoría creada con ID: {categoria_id}")
        
        # ============================================
        # 4. ACTUALIZACIÓN SOLO DEL INGRESO ACTUAL
        # ============================================
        print(f"\n🔄 ACTUALIZANDO INGRESO (solo este):")
        
        cambios = []
        
        # Solo comparar si el ID de categoría cambió
        if ingreso.categoria_id != categoria_id:
            print(f"   🏷️  Categoría ID: {ingreso.categoria_id} → {categoria_id}")
            
            # Obtener info de la categoría anterior y nueva para el log
            cat_vieja = db.query(models.Categoria).filter(
                models.Categoria.id == ingreso.categoria_id
            ).first() if ingreso.categoria_id else None
            
            cat_nueva = db.query(models.Categoria).filter(
                models.Categoria.id == categoria_id
            ).first()
            
            if cat_vieja and cat_nueva:
                print(f"   📝 Categoría: '{cat_vieja.nombre}' ({cat_vieja.tipo}) → '{cat_nueva.nombre}' ({cat_nueva.tipo})")
            elif cat_nueva:
                print(f"   📝 Nueva categoría: '{cat_nueva.nombre}' ({cat_nueva.tipo})")
            
            ingreso.categoria_id = categoria_id
            cambios.append("categoría")
        
        # Verificar otros campos
        if float(ingreso.valor) != float(valor):
            print(f"   💰 Valor: {ingreso.valor} → {valor}")
            ingreso.valor = valor
            cambios.append("valor")
        
        if ingreso.fecha != fecha_dt:
            print(f"   📅 Fecha: {ingreso.fecha} → {fecha_dt}")
            ingreso.fecha = fecha_dt
            cambios.append("fecha")
        
        if ingreso.estado != estado:
            print(f"   ✅ Estado: '{ingreso.estado}' → '{estado}'")
            ingreso.estado = estado
            cambios.append("estado")
        
        notas_actual = ingreso.notas if ingreso.notas else ""
        notas_nueva = notas if notas else ""
        
        if notas_actual.strip() != notas_nueva.strip():
            print(f"   📝 Notas: '{notas_actual}' → '{notas_nueva}'")
            ingreso.notas = notas
            cambios.append("notas")
        
        # ============================================
        # 5. GUARDAR CAMBIOS Y VERIFICAR
        # ============================================
        if not cambios:
            print(f"\nℹ️  No hay cambios detectados. Nada que actualizar.")
            mensaje_usuario = "ℹ️  No se realizaron cambios"
        else:
            print(f"\n✅ Cambios a realizar ({len(cambios)}): {', '.join(cambios)}")
            db.commit()
            print(f"💾 Cambios guardados en la base de datos")
            
            # Verificar que solo este ingreso se modificó
            print(f"\n📊 VERIFICACIÓN POST-COMMIT:")
            print(f"   Ingreso ID {id} actualizado correctamente")
            
            # Verificar que otras categorías no se modificaron
            if 'categoría' in cambios:
                # Contar cuántos ingresos usan la categoría vieja vs nueva
                ingresos_categoria_vieja = db.query(models.Ingreso).filter(
                    models.Ingreso.categoria_id == ingreso.categoria_id
                ).count() if ingreso.categoria_id else 0
                
                print(f"   {ingresos_categoria_vieja} ingresos usan ahora esta categoría")
            
            mensaje_usuario = f"✅ Ingreso actualizado. Campos modificados: {', '.join(cambios)}"
        
        # ============================================
        # 6. MENSAJE FINAL AL USUARIO
        # ============================================
        request.session['mensaje'] = {
            'tipo': 'exito',
            'titulo': '¡Éxito!',
            'texto': mensaje_usuario
        }
        
        print(f"\n🎯 {mensaje_usuario}")
        print(f"{'='*60}")
        
    except ValueError as e:
        print(f"\n❌ ERROR DE VALIDACIÓN: {str(e)}")
        db.rollback()
        request.session['mensaje'] = {
            'tipo': 'error',
            'titulo': 'Error de validación',
            'texto': f'Error: {str(e)}'
        }
        
    except HTTPException as e:
        print(f"\n❌ ERROR HTTP: {e.detail}")
        db.rollback()
        request.session['mensaje'] = {
            'tipo': 'error',
            'titulo': 'Error',
            'texto': f'Error: {e.detail}'
        }
        
    except Exception as e:
        print(f"\n❌ ERROR INESPERADO: {str(e)}")
        import traceback
        traceback.print_exc()
        db.rollback()
        request.session['mensaje'] = {
            'tipo': 'error',
            'titulo': 'Error',
            'texto': f'Error inesperado: {str(e)}'
        }

    return RedirectResponse(url="/ingresos", status_code=303)



@router.get("/ingresos/eliminar/{ingreso_id}")
def eliminar_ingreso(
    request: Request,
    ingreso_id: int,
    db: Session = Depends(get_db)
):
    """Eliminar un ingreso"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=303)

    ingreso = crud.obtener_ingreso(db, ingreso_id)
    if not ingreso or ingreso.usuario_id != usuario_id:
        raise HTTPException(status_code=404, detail="Ingreso no encontrado")

    crud.eliminar_ingreso(db, ingreso_id)
    request.session["mensaje"] = {
        "tipo": "exito",
        "titulo": "¡Eliminado!",
        "texto": "Ingreso eliminado correctamente"
    }
    return RedirectResponse(url="/ingresos", status_code=303)

# ============================================================================
# RUTAS PARA DEBUG Y MANTENIMIENTO
# ============================================================================

@router.get("/debug/ingresos-completo")
def debug_ingresos_completo(
    request: Request,
    db: Session = Depends(get_db)
):
    """Endpoint para debug: mostrar todos los ingresos"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return {"error": "No autenticado"}
    
    # Todos los ingresos
    todos_ingresos = db.query(
        models.Ingreso.id,
        models.Ingreso.usuario_id,
        models.Ingreso.estado,
        models.Ingreso.valor,
        models.Ingreso.fecha,
        models.Categoria.nombre.label('categoria_nombre'),
        models.Categoria.tipo.label('categoria_tipo')
    ).join(
        models.Categoria, models.Ingreso.categoria_id == models.Categoria.id
    ).all()
    
    # Ingresos del usuario actual
    ingresos_usuario = db.query(
        models.Ingreso.id,
        models.Ingreso.estado,
        models.Ingreso.valor,
        models.Ingreso.fecha,
        models.Categoria.nombre.label('categoria_nombre'),
        models.Categoria.tipo.label('categoria_tipo')
    ).join(
        models.Categoria, models.Ingreso.categoria_id == models.Categoria.id
    ).filter(models.Ingreso.usuario_id == usuario_id).all()
    
    return {
        "usuario_actual": usuario_id,
        "todos_ingresos": [
            {
                "id": row.id,
                "usuario_id": row.usuario_id,
                "estado": row.estado,
                "valor": float(row.valor),
                "fecha": str(row.fecha),
                "categoria": row.categoria_nombre,
                "tipo": row.categoria_tipo
            }
            for row in todos_ingresos
        ],
        "ingresos_usuario_actual": [
            {
                "id": row.id,
                "estado": row.estado,
                "valor": float(row.valor),
                "fecha": str(row.fecha),
                "categoria": row.categoria_nombre,
                "tipo": row.categoria_tipo
            }
            for row in ingresos_usuario
        ],
        "resumen": {
            "total_ingresos": len(todos_ingresos),
            "ingresos_usuario_actual": len(ingresos_usuario),
            "estados_usuario_actual": {
                "pendiente": sum(1 for row in ingresos_usuario if row.estado == 'pendiente'),
                "recibido": sum(1 for row in ingresos_usuario if row.estado == 'recibido')
            }
        }
    }

@router.get("/reparar/ingresos")
def reparar_ingresos_route(
    request: Request,
    db: Session = Depends(get_db)
):
    """Reparar ingresos corruptos (sin categoría)"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return {"error": "No autenticado"}
    
    problemas = crud.reparar_ingresos_corruptos(db, usuario_id)
    
    return {
        "status": "success",
        "problemas_reparados": problemas,
        "message": f"Se repararon {problemas} ingresos con categorías faltantes"
    }

# ============================================================================
# RUTAS DE GASTOS (MANTENIDAS)
# ============================================================================

@router.get("/gastos")
def listar_gastos(
    request: Request,
    db: Session = Depends(get_db),
    page: int = 1,
    tipo: Optional[str] = None,
    pagado: Optional[str] = None  # Cambia de bool a Optional[str]
):
    """Listar todos los gastos del usuario"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=303)
    
    print(f"\n=== LISTANDO GASTOS PARA USUARIO ID: {usuario_id} ===")
    print(f"Filtros recibidos - tipo: '{tipo}', pagado: '{pagado}'")
    
    # Convertir parámetro 'pagado' de string a booleano o None
    pagado_bool = None
    if pagado is not None and pagado != "":
        if pagado.lower() == 'true':
            pagado_bool = True
        elif pagado.lower() == 'false':
            pagado_bool = False
        else:
            print(f"⚠️  Valor inválido para 'pagado': '{pagado}', ignorando filtro")
    
    print(f"Filtro 'pagado' convertido a booleano: {pagado_bool}")
    
    data = crud.obtener_gastos_paginados(
        db,
        usuario_id,
        page=page,
        tipo=tipo,
        pagado=pagado_bool  # Pasar el booleano convertido
    )
    
    mensaje = request.session.pop("mensaje", None)
    
    return templates.TemplateResponse(
        "gastos_listado.html",
        {
            "request": request,
            "gastos": data["gastos"],
            "total_pages": data["total_pages"],
            "current_page": page,
            "filtro_tipo": tipo,
            "filtro_pagado": pagado,  # Mantener el string original para el template
            "mensaje": mensaje
        }
    )

@router.get("/gastos/nuevo", response_class=HTMLResponse)
@router.get("/gastos/editar/{gasto_id}", response_class=HTMLResponse)
def formulario_gasto(
    request: Request,
    gasto_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """Mostrar formulario para crear/editar gasto"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=303)

    gasto = None
    if gasto_id:
        gasto = crud.obtener_gasto(db, gasto_id)
        if not gasto or gasto.usuario_id != usuario_id:
            raise HTTPException(status_code=404, detail="Gasto no encontrado")

    categorias = crud.obtener_categorias(db, usuario_id)
    
    return templates.TemplateResponse("gasto_form.html", {
        "request": request,
        "gasto": gasto,
        "categorias": categorias
    })

@router.post("/gastos/guardar")
def guardar_gasto(
    request: Request,
    id: Optional[int] = Form(None),
    categoria_nombre: str = Form(...),
    tipo_categoria: str = Form("variable"),
    valor: float = Form(...),
    fecha_limite: Optional[str] = Form(None),
    pagado: bool = Form(False),
    notas: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    """Guardar gasto (crear o actualizar)"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=303)

    try:
        print(f"\n{'='*60}")
        print(f"{'📝 EDITANDO GASTO' if id else '➕ CREANDO NUEVO GASTO'}")
        print(f"{'='*60}")
        
        # Normalizar datos
        categoria_nombre_clean = categoria_nombre.strip()
        
        # Convertir fecha si existe
        fecha_limite_dt = None
        if fecha_limite and fecha_limite.strip():
            fecha_limite_dt = date.fromisoformat(fecha_limite.strip())
        
        # Buscar o crear categoría (similar a ingresos)
        # Buscar por nombre + tipo
        categoria_existente = db.query(models.Categoria).filter(
            models.Categoria.nombre == categoria_nombre_clean,
            models.Categoria.tipo == tipo_categoria,
            models.Categoria.usuario_id == usuario_id
        ).first()
        
        if categoria_existente:
            print(f"✅ Usando categoría existente: '{categoria_existente.nombre}' ({categoria_existente.tipo})")
            categoria_id = categoria_existente.id
        else:
            print(f"➕ Creando nueva categoría: '{categoria_nombre_clean}' ({tipo_categoria})")
            nueva_categoria = models.Categoria(
                nombre=categoria_nombre_clean,
                tipo=tipo_categoria,
                usuario_id=usuario_id
            )
            db.add(nueva_categoria)
            db.flush()
            categoria_id = nueva_categoria.id
        
        # Crear datos del gasto
        gasto_data = schemas.GastoCreate(
            categoria_id=categoria_id,
            valor=valor,
            fecha_limite=fecha_limite_dt,
            pagado=pagado,
            notas=notas
        )

        if id:
            # Actualizar gasto existente
            crud.actualizar_gasto(db, gasto_id=int(id), gasto=gasto_data, usuario_id=usuario_id)
            mensaje = "Gasto actualizado correctamente"
        else:
            # Crear nuevo gasto
            crud.crear_gasto(db, gasto=gasto_data, usuario_id=usuario_id)
            mensaje = "Gasto creado correctamente"

        request.session['mensaje'] = {
            'tipo': 'exito',
            'titulo': '¡Éxito!',
            'texto': mensaje
        }
        
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        db.rollback()
        request.session['mensaje'] = {
            'tipo': 'error',
            'titulo': 'Error',
            'texto': f'Ocurrió un error: {str(e)}'
        }

    return RedirectResponse(url="/gastos", status_code=303)

@router.get("/gastos/eliminar/{gasto_id}")
def eliminar_gasto(
    request: Request,
    gasto_id: int,
    db: Session = Depends(get_db)
):
    """Eliminar un gasto"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=303)

    gasto = crud.obtener_gasto(db, gasto_id)
    if not gasto or gasto.usuario_id != usuario_id:
        raise HTTPException(status_code=404, detail="Gasto no encontrado")

    crud.eliminar_gasto(db, gasto_id)
    request.session["mensaje"] = {
        "tipo": "exito",
        "titulo": "¡Eliminado!",
        "texto": "Gasto eliminado correctamente"
    }
    return RedirectResponse(url="/gastos", status_code=303)

# ============================================================================
# RUTAS DE PENDIENTES
# ============================================================================

@router.get("/pendientes", response_class=HTMLResponse)
async def listar_pendientes(
    request: Request,
    estado: Optional[str] = None,
    prioridad: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Listar pendientes del usuario"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)

    pendientes = crud.get_pendientes_by_filters(
        db,
        usuario_id=usuario_id,
        estado=estado,
        prioridad=prioridad
    )

    ahora = datetime.now()
    recordatorios_vencidos = [
        p for p in pendientes if p.recordatorio and p.recordatorio <= ahora
    ]

    return templates.TemplateResponse(
        "pendientes.html",
        {
            "request": request,
            "pendientes": pendientes,
            "recordatorios_vencidos": recordatorios_vencidos,
            "now": ahora
        }
    )

@router.get("/pendientes/nuevo", response_class=HTMLResponse)
@router.get("/pendientes/editar/{pendiente_id}", response_class=HTMLResponse)
async def form_pendiente(
    request: Request,
    pendiente_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """Mostrar formulario para crear/editar pendiente"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)

    pendiente = None
    if pendiente_id:
        pendiente = crud.get_pendiente(db, pendiente_id)
        if not pendiente or pendiente.usuario_id != usuario_id:
            raise HTTPException(status_code=404, detail="Pendiente no encontrado")

    return templates.TemplateResponse(
        "pendientes_form.html",
        {
            "request": request,
            "pendiente": pendiente
        }
    )

@router.post("/pendientes/guardar")
async def guardar_pendiente(
    request: Request,
    id: Optional[int] = Form(None),
    titulo: str = Form(...),
    descripcion: Optional[str] = Form(None),
    estado: str = Form(...),
    prioridad: str = Form(...),
    fecha_limite: Optional[str] = Form(None),
    recordatorio: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    """Guardar pendiente (crear o actualizar)"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)

    fecha_limite_dt = datetime.fromisoformat(fecha_limite) if fecha_limite else None
    recordatorio_dt = datetime.fromisoformat(recordatorio) if recordatorio else None

    pendiente_data = schemas.PendienteCreate(
        titulo=titulo,
        descripcion=descripcion,
        estado=estado,
        prioridad=prioridad,
        fecha_limite=fecha_limite_dt,
        recordatorio=recordatorio_dt
    )

    try:
        if id:
            pendiente = crud.update_pendiente(db, pendiente_id=id, pendiente=pendiente_data)
            mensaje = "Pendiente actualizado correctamente"
        else:
            pendiente = crud.create_pendiente(db, pendiente=pendiente_data, usuario_id=usuario_id)
            mensaje = "Pendiente creado correctamente"

        request.session['mensaje'] = {
            'tipo': 'exito',
            'titulo': '¡Éxito!',
            'texto': mensaje
        }
    except Exception as e:
        request.session['mensaje'] = {
            'tipo': 'error',
            'titulo': 'Error',
            'texto': f'Ocurrió un error: {str(e)}'
        }

    return RedirectResponse(url="/pendientes", status_code=HTTP_303_SEE_OTHER)

@router.get("/pendientes/eliminar/{pendiente_id}")
async def eliminar_pendiente(
    request: Request,
    pendiente_id: int,
    db: Session = Depends(get_db)
):
    """Eliminar un pendiente"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)

    pendiente = crud.get_pendiente(db, pendiente_id)
    if not pendiente or pendiente.usuario_id != usuario_id:
        raise HTTPException(status_code=404, detail="Pendiente no encontrado")

    crud.delete_pendiente(db, pendiente_id)
    request.session["mensaje"] = {
        "tipo": "exito",
        "titulo": "¡Eliminado!",
        "texto": "Pendiente eliminado correctamente"
    }
    return RedirectResponse(url="/pendientes", status_code=HTTP_303_SEE_OTHER)

# ============================================================================
# RUTAS DE CONTRASEÑAS
# ============================================================================

@router.get("/contrasenas", response_class=HTMLResponse)
def listar_contrasenas(
    request: Request,
    db: Session = Depends(get_db),
    page: int = 1,
    items_per_page: int = 10
):
    """Listar contraseñas del usuario"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=303)
    
    contrasenas = crud.obtener_contrasenas_usuario(db, usuario_id)
    
    # Calcular paginación
    total_items = len(contrasenas)
    total_pages = (total_items + items_per_page - 1) // items_per_page if total_items > 0 else 1
    
    # Ajustar página
    if page < 1:
        page = 1
    elif page > total_pages and total_pages > 0:
        page = total_pages
    
    # Obtener elementos para la página actual
    start_idx = (page - 1) * items_per_page
    end_idx = start_idx + items_per_page
    contrasenas_pagina = contrasenas[start_idx:end_idx]
    
    mensaje = request.session.pop("mensaje", None)
    
    return templates.TemplateResponse(
        "contrasenas_listado.html",
        {
            "request": request,
            "contrasenas": contrasenas_pagina,
            "total_pages": total_pages,
            "current_page": page,
            "items_per_page": items_per_page,
            "total_items": total_items,
            "mensaje": mensaje
        }
    )

@router.get("/contrasenas/nuevo", response_class=HTMLResponse)
@router.get("/contrasenas/editar/{contrasena_id}", response_class=HTMLResponse)
def formulario_contrasena(
    request: Request,
    contrasena_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """Mostrar formulario para crear/editar contraseña"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=303)
    
    contrasena = None
    if contrasena_id:
        contrasena = crud.obtener_contrasena(db, contrasena_id)
        if not contrasena or contrasena.usuario_id != usuario_id:
            raise HTTPException(status_code=404, detail="Contraseña no encontrada")
    
    return templates.TemplateResponse("contrasenas_form.html", {
        "request": request,
        "contrasena": contrasena
    })

@router.post("/contrasenas/guardar")
def guardar_contrasena(
    request: Request,
    id: Optional[int] = Form(None),
    servicio: str = Form(...),
    usuario: str = Form(...),
    contrasena: str = Form(...),
    url: Optional[str] = Form(None),
    notas: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    """Guardar contraseña (crear o actualizar)"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=303)
    
    contrasena_data = schemas.ContrasenaCreate(
        servicio=servicio,
        usuario=usuario,
        contrasena=contrasena,
        url=url,
        notas=notas
    )
    
    try:
        if id:
            contrasena_update = schemas.ContrasenaUpdate(
                servicio=servicio,
                usuario=usuario,
                contrasena=contrasena,
                url=url,
                notas=notas
            )
            crud.actualizar_contrasena(db, contrasena_id=int(id), contrasena=contrasena_update, usuario_id=usuario_id)
            mensaje = "Contraseña actualizada correctamente"
        else:
            crud.crear_contrasena(db, contrasena=contrasena_data, usuario_id=usuario_id)
            mensaje = "Contraseña creada correctamente"
        
        request.session['mensaje'] = {
            'tipo': 'exito',
            'titulo': '¡Éxito!',
            'texto': mensaje
        }
    except Exception as e:
        request.session['mensaje'] = {
            'tipo': 'error',
            'titulo': 'Error',
            'texto': f'Ocurrió un error: {str(e)}'
        }
    
    return RedirectResponse(url="/contrasenas", status_code=303)

@router.get("/contrasenas/eliminar/{contrasena_id}")
def eliminar_contrasena(
    request: Request,
    contrasena_id: int,
    db: Session = Depends(get_db)
):
    """Eliminar una contraseña"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=303)
    
    contrasena = crud.obtener_contrasena(db, contrasena_id)
    if not contrasena or contrasena.usuario_id != usuario_id:
        raise HTTPException(status_code=404, detail="Contraseña no encontrada")
    
    crud.eliminar_contrasena(db, contrasena_id, usuario_id)
    
    request.session["mensaje"] = {
        "tipo": "exito",
        "titulo": "¡Eliminado!",
        "texto": "Contraseña eliminada correctamente"
    }
    
    return RedirectResponse(url="/contrasenas", status_code=303)


@router.get("/contrasenas/obtener/{contrasena_id}")
def obtener_contrasena_desencriptada(
    request: Request,
    contrasena_id: int,
    db: Session = Depends(get_db)
):
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return JSONResponse(content={"status": "error", "mensaje": "No autenticado"}, status_code=401)
    
    # Obtener la contraseña desencriptada
    contrasena_texto = crud.desencriptar_contrasena_db(db, contrasena_id, usuario_id)
    
    if not contrasena_texto:
        return JSONResponse(content={"status": "error", "mensaje": "Contraseña no encontrada"}, status_code=404)
    
    return JSONResponse(content={
        "status": "success",
        "contrasena": contrasena_texto
    })



# ============================================================================
# RUTAS DE CUMPLEAÑOS
# ============================================================================

@router.get("/cumpleanos", response_class=HTMLResponse)
async def listar_cumpleanos(
    request: Request,
    page: int = 1,
    relacion: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Listar cumpleaños"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse("/", status_code=303)
    
    resultado = crud.obtener_cumpleanos_paginados(
        db,
        usuario_id=usuario_id,
        page=page,
        per_page=10,
        relacion=relacion
    )
    
    # Calcular días, edad y próximo cumpleaños
    for cumple in resultado["cumpleanos"]:
        cumple.dias_hasta_cumpleanos = crud.calcular_dias_hasta_cumpleanos(cumple.fecha_nacimiento)
        cumple.edad = crud.calcular_edad(cumple.fecha_nacimiento)
        cumple.proximo_cumple = crud.calcular_proximo_cumple(cumple.fecha_nacimiento)

    return templates.TemplateResponse("cumpleanos_listado.html", {
        "request": request,
        "cumpleanos": resultado["cumpleanos"],
        "current_page": resultado["current_page"],
        "total_pages": resultado["total_pages"],
        "filtro_relacion": relacion
    })

@router.get("/cumpleanos/nuevo", response_class=HTMLResponse)
async def formulario_nuevo_cumpleano(request: Request):
    """Mostrar formulario para nuevo cumpleaño"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse("/", status_code=303)
    
    return templates.TemplateResponse("cumpleanos_form.html", {
        "request": request,
        "cumpleano": None
    })

@router.get("/cumpleanos/editar/{cumpleano_id}", response_class=HTMLResponse)
async def formulario_editar_cumpleano(
    request: Request,
    cumpleano_id: int,
    db: Session = Depends(get_db)
):
    """Mostrar formulario para editar cumpleaño"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse("/", status_code=303)
    
    cumpleano = crud.obtener_cumpleano(db, cumpleano_id)
    if not cumpleano or cumpleano.usuario_id != usuario_id:
        return RedirectResponse("/cumpleanos", status_code=303)
    
    return templates.TemplateResponse("cumpleanos_form.html", {
        "request": request,
        "cumpleano": cumpleano
    })

@router.post("/cumpleanos/guardar")
async def guardar_cumpleano(
    request: Request,
    db: Session = Depends(get_db)
):
    """Guardar cumpleaño (crear o actualizar)"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse("/", status_code=303)
    
    form_data = await request.form()
    cumpleano_id = form_data.get("id")
    
    cumpleano_data = schemas.CumpleanoCreate(
        nombre_persona=form_data.get("nombre_persona"),
        fecha_nacimiento=datetime.strptime(form_data.get("fecha_nacimiento"), "%Y-%m-%d").date(),
        telefono=form_data.get("telefono") or None,
        email=form_data.get("email") or None,
        relacion=form_data.get("relacion") or None,
        notas=form_data.get("notas") or None,
        notificar_dias_antes=int(form_data.get("notificar_dias_antes", 7))
    )
    
    if cumpleano_id:
        # Actualizar
        crud.actualizar_cumpleano(db, int(cumpleano_id), schemas.CumpleanoUpdate(**cumpleano_data.model_dump()), usuario_id)
    else:
        # Crear nuevo
        crud.crear_cumpleano(db, cumpleano_data, usuario_id)
    
    return RedirectResponse("/cumpleanos", status_code=303)

@router.get("/cumpleanos/eliminar/{cumpleano_id}")
async def eliminar_cumpleano(
    request: Request,
    cumpleano_id: int,
    db: Session = Depends(get_db)
):
    """Eliminar un cumpleaño"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse("/", status_code=303)
    
    crud.eliminar_cumpleano(db, cumpleano_id, usuario_id)
    return RedirectResponse("/cumpleanos", status_code=303)

# ============================================================================
# RUTA EXTRA PARA DEBUG
# ============================================================================

@router.post("/debug/form")
async def debug_form(request: Request):
    """
    Endpoint para debug: mostrar todos los parámetros recibidos del formulario
    Útil para verificar qué datos está enviando el formulario
    """
    form_data = await request.form()
    print("\n📋 DEBUG - DATOS RECIBIDOS DEL FORMULARIO:")
    for key, value in form_data.items():
        print(f"  {key}: {value}")
    
    return {"status": "ok", "data": dict(form_data)}


# ============================================================================
# 💳 FUNCIONES DE CRÉDITOS
# ============================================================================

@router.post("/creditos/guardar")
def guardar_credito(
    request: Request,
    id: Optional[int] = Form(None),
    nombre_credito: str = Form(...),
    monto: float = Form(...),
    interes: float = Form(...),
    plazo_meses: int = Form(...),
    frecuencia_pago: str = Form("mensual"),
    fecha_inicio: str = Form(...),
    seguro: float = Form(0),
    cuota_manual: float = Form(0),  # ✅ AÑADIDO
    observaciones: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    """Guardar crédito (crear o actualizar)"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        print(f"\n{'='*60}")
        print(f"{'📝 EDITANDO CRÉDITO' if id else '➕ CREANDO NUEVO CRÉDITO'}")
        print(f"Cuota manual recibida: ${cuota_manual:,.2f}")
        print(f"{'='*60}")
        
        fecha_inicio_dt = date.fromisoformat(fecha_inicio)
        
        credito_data = schemas.CreditoCreate(
            nombre_credito=nombre_credito,
            monto=monto,
            interes=interes,
            plazo_meses=plazo_meses,
            frecuencia_pago=frecuencia_pago,
            fecha_inicio=fecha_inicio_dt,
            seguro=seguro,
            cuota_manual=cuota_manual,  # ✅ Pasar la cuota manual
            observaciones=observaciones
        )
        
        if id:
            credito_update = schemas.CreditoUpdate(**credito_data.model_dump())
            crud.actualizar_credito(db, credito_id=int(id), credito=credito_update, usuario_id=usuario_id)
            mensaje = "Crédito actualizado correctamente"
        else:
            crud.crear_credito(db, credito=credito_data, usuario_id=usuario_id)
            mensaje = "Crédito creado correctamente"
        
        request.session['mensaje'] = {
            'tipo': 'exito',
            'titulo': '¡Éxito!',
            'texto': mensaje
        }
        
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        db.rollback()
        request.session['mensaje'] = {
            'tipo': 'error',
            'titulo': 'Error',
            'texto': f'Error: {str(e)}'
        }
    
    return RedirectResponse(url="/creditos", status_code=303)


def calcular_cuota_credito(monto: float, interes_mensual: float, plazo_meses: int, 
                          frecuencia: str = 'mensual') -> float:
    """
    Calcula la cuota de un crédito usando fórmula de amortización francesa
    
    Args:
        monto: Monto del crédito
        interes_mensual: Tasa de interés MENSUAL en porcentaje (ej: 1.4 para 1.4%)
        plazo_meses: Plazo en meses
        frecuencia: Frecuencia de pago ('mensual' principalmente)
    
    Returns:
        Cuota calculada (solo del crédito, sin seguro)
    """
    if interes_mensual == 0:
        return round(monto / plazo_meses, 2)
    
    i = interes_mensual / 100
    
    if frecuencia == 'quincenal':
        i = i / 2
        n = plazo_meses * 2
    elif frecuencia == 'semanal':
        i = i / 4
        n = plazo_meses * 4
    elif frecuencia == 'diario':
        i = i / 30
        n = plazo_meses * 30
    else:
        n = plazo_meses
    
    try:
        numerador = i * pow(1 + i, n)
        denominador = pow(1 + i, n) - 1
        cuota = monto * (numerador / denominador)
        return round(cuota, 2)
    except Exception as e:
        print(f"Error en cálculo de cuota: {e}")
        return 0.0

def crear_credito(db: Session, credito: schemas.CreditoCreate, usuario_id: int):
    """Crea un nuevo crédito para un usuario con dos modos: cálculo o manual"""
    try:
        # 🔹 MODE 1: CUOTA MANUAL (BANCO)
        if credito.cuota_manual and credito.cuota_manual > 0:
            cuota_real = credito.cuota_manual
            cuota_calculada = calcular_cuota_credito(
                credito.monto, credito.interes, credito.plazo_meses, credito.frecuencia_pago
            )
            print(f"🏦 MODO MANUAL ACTIVADO")
            print(f"   Cuota banco: ${cuota_real:,.2f}")
            print(f"   Cuota calculada: ${cuota_calculada:,.2f}")
            print(f"   Diferencia: ${cuota_real - cuota_calculada:,.2f}")
        
        # 🔹 MODE 2: CÁLCULO AUTOMÁTICO
        else:
            cuota_calculada = calcular_cuota_credito(
                credito.monto, credito.interes, credito.plazo_meses, credito.frecuencia_pago
            )
            cuota_real = cuota_calculada  # Son iguales en modo cálculo
            print(f"📐 MODO CÁLCULO AUTOMÁTICO")
            print(f"   Cuota calculada: ${cuota_calculada:,.2f}")
        
        print(f"\n💳 RESUMEN CRÉDITO:")
        print(f"   Monto: ${credito.monto:,.2f}")
        print(f"   Interés: {credito.interes}% mensual")
        print(f"   Plazo: {credito.plazo_meses} meses")
        print(f"   Seguro: ${credito.seguro:,.2f}")
        print(f"   Cuota REAL a pagar: ${cuota_real:,.2f}")
        print(f"   Cuota TOTAL (con seguro): ${cuota_real + credito.seguro:,.2f}")
        
        # Calcular total a pagar (cuota REAL + seguro) * número de cuotas
        if credito.frecuencia_pago == 'quincenal':
            total_cuotas = credito.plazo_meses * 2
        elif credito.frecuencia_pago == 'semanal':
            total_cuotas = credito.plazo_meses * 4
        elif credito.frecuencia_pago == 'diario':
            total_cuotas = credito.plazo_meses * 30
        else:
            total_cuotas = credito.plazo_meses
        
        total_pagar = (cuota_real + credito.seguro) * total_cuotas
        
        # Crear crédito con AMBAS cuotas
        db_credito = models.Credito(
            nombre_credito=credito.nombre_credito,
            monto=credito.monto,
            interes=credito.interes,
            plazo_meses=credito.plazo_meses,
            frecuencia_pago=credito.frecuencia_pago,
            fecha_inicio=credito.fecha_inicio,
            
            # ✅ GUARDAR AMBAS
            cuota_manual=credito.cuota_manual,  # Lo que ingresó el usuario (0 o valor)
            cuota=cuota_real,  # Lo que REALMENTE se pagará (manual o calculada)
            cuota_calculada=calcular_cuota_credito(  # Lo que la fórmula dice
                credito.monto, credito.interes, credito.plazo_meses, credito.frecuencia_pago
            ),
            
            seguro=credito.seguro,
            total_pagar=total_pagar,
            saldo_actual=credito.monto,
            estado='activo',
            observaciones=credito.observaciones,
            usuario_id=usuario_id
        )
        
        db.add(db_credito)
        db.commit()
        db.refresh(db_credito)
        
        print(f"✅ Crédito creado con ID: {db_credito.id}")
        print(f"   Modo: {'MANUAL (banco)' if credito.cuota_manual > 0 else 'CÁLCULO automático'}")
        
        return db_credito
        
    except Exception as e:
        print(f"❌ Error al crear crédito: {e}")
        import traceback
        traceback.print_exc()
        db.rollback()
        return None

def obtener_credito(db: Session, credito_id: int):
    return db.query(models.Credito).filter(models.Credito.id == credito_id).first()

def obtener_creditos_paginados(db: Session, usuario_id: int, page: int = 1,
                               page_size: int = 10, estado: Optional[str] = None,
                               frecuencia: Optional[str] = None):
    query = db.query(models.Credito).filter(models.Credito.usuario_id == usuario_id)
    
    if estado:
        query = query.filter(models.Credito.estado == estado)
    if frecuencia:
        query = query.filter(models.Credito.frecuencia_pago == frecuencia)
    
    total_items = query.count()
    total_pages = (total_items + page_size - 1) // page_size if total_items > 0 else 1
    
    creditos = query.order_by(models.Credito.fecha_inicio.desc()) \
                    .offset((page - 1) * page_size) \
                    .limit(page_size) \
                    .all()
    
    return {
        "creditos": creditos,
        "total_pages": total_pages,
        "current_page": page
    }

def actualizar_credito(db: Session, credito_id: int, credito: schemas.CreditoUpdate, usuario_id: int):
    """Actualiza un crédito existente - VERSIÓN CORREGIDA"""
    db_credito = obtener_credito(db, credito_id)
    if not db_credito or db_credito.usuario_id != usuario_id:
        return None
    
    # ✅ CORREGIDO: Inicializar variables
    recalcular_total = False
    cuota_a_usar = db_credito.cuota
    ha_cambiado_cuota_manual = False
    
    # 🔹 1. Si cambia la cuota manual
    if credito.cuota_manual is not None:
        db_credito.cuota_manual = credito.cuota_manual
        ha_cambiado_cuota_manual = True
        
        if credito.cuota_manual > 0:
            # MODO MANUAL: Usar valor del banco
            cuota_a_usar = credito.cuota_manual
            print(f"🏦 Actualizando a CUOTA MANUAL: ${cuota_a_usar:,.2f}")
        else:
            # MODO CÁLCULO: Volver a calcular
            cuota_a_usar = calcular_cuota_credito(
                db_credito.monto, db_credito.interes, 
                db_credito.plazo_meses, db_credito.frecuencia_pago
            )
            print(f"📐 Volviendo a CÁLCULO AUTOMÁTICO: ${cuota_a_usar:,.2f}")
        
        recalcular_total = True
    
    # 🔹 2. Si cambian parámetros de cálculo (monto, interés, plazo, frecuencia)
    parametros_modificados = any([
        credito.monto is not None,
        credito.interes is not None,
        credito.plazo_meses is not None,
        credito.frecuencia_pago is not None
    ])
    
    if parametros_modificados and not ha_cambiado_cuota_manual:
        # Determinar valores nuevos o mantener actuales
        monto = credito.monto if credito.monto is not None else db_credito.monto
        interes = credito.interes if credito.interes is not None else db_credito.interes
        plazo = credito.plazo_meses if credito.plazo_meses is not None else db_credito.plazo_meses
        frecuencia = credito.frecuencia_pago if credito.frecuencia_pago is not None else db_credito.frecuencia_pago
        
        # ¿Usar cuota manual existente o recalcular?
        if db_credito.cuota_manual and db_credito.cuota_manual > 0:
            # Mantener cuota manual (no recalcular aunque cambien parámetros)
            cuota_a_usar = db_credito.cuota_manual
            print(f"📌 Manteniendo CUOTA MANUAL existente: ${cuota_a_usar:,.2f}")
        else:
            # Calcular nueva cuota
            cuota_a_usar = calcular_cuota_credito(monto, interes, plazo, frecuencia)
            print(f"🔄 Recalculando cuota: ${cuota_a_usar:,.2f}")
        
        recalcular_total = True
    
    # 🔹 3. Actualizar la cuota REAL que se usará
    db_credito.cuota = cuota_a_usar
    
    # 🔹 4. Recalcular total_pagar si es necesario
    if recalcular_total or credito.seguro is not None:
        # Determinar número de cuotas según frecuencia
        frecuencia_actual = credito.frecuencia_pago if credito.frecuencia_pago is not None else db_credito.frecuencia_pago
        plazo_actual = credito.plazo_meses if credito.plazo_meses is not None else db_credito.plazo_meses
        
        if frecuencia_actual == 'quincenal':
            total_cuotas = plazo_actual * 2
        elif frecuencia_actual == 'semanal':
            total_cuotas = plazo_actual * 4
        elif frecuencia_actual == 'diario':
            total_cuotas = plazo_actual * 30
        else:
            total_cuotas = plazo_actual
        
        # Determinar seguro actual
        seguro_actual = credito.seguro if credito.seguro is not None else db_credito.seguro
        
        # Calcular total a pagar
        db_credito.total_pagar = (cuota_a_usar + seguro_actual) * total_cuotas
    
    # 🔹 5. Actualizar otros campos
    for key, value in credito.model_dump(exclude_unset=True).items():
        if key not in ['cuota', 'total_pagar']:  # Estos ya los manejamos
            setattr(db_credito, key, value)
    
    # 🔹 6. Recalcular cuota_calculada (siempre)
    db_credito.cuota_calculada = calcular_cuota_credito(
        db_credito.monto, db_credito.interes, 
        db_credito.plazo_meses, db_credito.frecuencia_pago
    )
    
    db.commit()
    db.refresh(db_credito)
    
    print(f"✅ Crédito actualizado")
    print(f"   Cuota real: ${db_credito.cuota:,.2f}")
    print(f"   Cuota calculada: ${db_credito.cuota_calculada:,.2f}")
    print(f"   Diferencia: ${db_credito.cuota - db_credito.cuota_calculada:,.2f}")
    
    return db_credito

# def eliminar_credito(db: Session, credito_id: int, usuario_id: int):
#     db_credito = obtener_credito(db, credito_id)
#     if not db_credito or db_credito.usuario_id != usuario_id:
#         return False
    
#     db.delete(db_credito)
#     db.commit()
#     return True


# ============================================================================
# RUTAS DE CRÉDITOS (COMPLETAS)
# ============================================================================

@router.get("/creditos")
def listar_creditos(
    request: Request,
    db: Session = Depends(get_db),
    page: int = 1,
    estado: Optional[str] = None,
    frecuencia: Optional[str] = None
):
    """Listar créditos del usuario"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=303)
    
    print(f"\n=== LISTANDO CRÉDITOS PARA USUARIO ID: {usuario_id} ===")
    print(f"Filtros - estado: '{estado}', frecuencia: '{frecuencia}'")
    
    data = crud.obtener_creditos_paginados(
        db,
        usuario_id,
        page=page,
        estado=estado,
        frecuencia=frecuencia
    )
    
    mensaje = request.session.pop("mensaje", None)
    
    return templates.TemplateResponse(
        "creditos_listado.html",
        {
            "request": request,
            "creditos": data["creditos"],
            "total_pages": data["total_pages"],
            "current_page": page,
            "filtro_estado": estado,
            "filtro_frecuencia": frecuencia,
            "mensaje": mensaje
        }
    )

@router.get("/creditos/nuevo", response_class=HTMLResponse)
@router.get("/creditos/editar/{credito_id}", response_class=HTMLResponse)
def formulario_credito(
    request: Request,
    credito_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """Mostrar formulario para crear/editar crédito"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=303)
    
    credito = None
    if credito_id:
        credito = crud.obtener_credito(db, credito_id)
        if not credito or credito.usuario_id != usuario_id:
            raise HTTPException(status_code=404, detail="Crédito no encontrado")
    
    return templates.TemplateResponse("creditos_form.html", {
        "request": request,
        "credito": credito
    })

@router.get("/creditos/eliminar/{credito_id}")
def eliminar_credito(
    request: Request,
    credito_id: int,
    db: Session = Depends(get_db)
):
    """Eliminar un crédito"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=303)
    
    credito = crud.obtener_credito(db, credito_id)
    if not credito or credito.usuario_id != usuario_id:
        raise HTTPException(status_code=404, detail="Crédito no encontrado")
    
    crud.eliminar_credito(db, credito_id, usuario_id)
    
    request.session["mensaje"] = {
        "tipo": "exito",
        "titulo": "¡Eliminado!",
        "texto": "Crédito eliminado correctamente"
    }
    
    return RedirectResponse(url="/creditos", status_code=303)

# ============================================================================
# 🔍 DETALLE DEL CRÉDITO (VERSIÓN MEJORADA)
# ============================================================================

# ============================================================================
# 🔍 DETALLE DEL CRÉDITO (VERSIÓN CORREGIDA)
# ============================================================================

@router.get("/creditos/detalle/{credito_id}")
def detalle_credito(
    request: Request,
    credito_id: int,
    db: Session = Depends(get_db)
):
    """Ver detalle completo de un crédito con su historial de pagos"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=303)
    
    # Obtener el crédito
    credito = crud.obtener_credito(db, credito_id)
    if not credito or credito.usuario_id != usuario_id:
        # ✅ CORREGIDO: Diccionario completo, no {...}
        request.session['mensaje'] = {
            'tipo': 'error',
            'titulo': 'Error',
            'texto': 'Crédito no encontrado'
        }
        return RedirectResponse(url="/creditos", status_code=303)
    
    # Obtener pagos del crédito
    pagos = crud.obtener_pagos_por_credito(db, credito_id)
    
    # Calcular total pagado
    total_pagado = sum(float(pago.monto) for pago in pagos) if pagos else 0
    
    # Calcular progreso
    progreso = (total_pagado / credito.monto * 100) if credito.monto > 0 else 0
    
    # Calcular cuota total
    cuota_total = credito.cuota + credito.seguro
    
    print(f"\n=== DETALLE CRÉDITO ID: {credito_id} ===")
    print(f"Usuario: {usuario_id}")
    print(f"Crédito: {credito.nombre_credito}")
    print(f"Monto: ${credito.monto:,.0f}")
    print(f"Pagos encontrados: {len(pagos)}")
    print(f"Total pagado: ${total_pagado:,.0f}")
    print(f"Progreso: {progreso:.1f}%")
    
    return templates.TemplateResponse(
        "credito_detalle.html",
        {
            "request": request,
            "credito": credito,
            "pagos": pagos,
            "total_pagado": total_pagado,
            "progreso": round(progreso, 1),
            "cuota_total": cuota_total,
            "mensaje": request.session.pop("mensaje", None)  # ✅ Agregar mensajes de sesión
        }
    )


# Agregar al archivo routes.py después de las rutas de cumpleaños

# ============================================================================
# RUTAS DE CONTACTOS
# ============================================================================

@router.get("/contactos")
def listar_contactos(
    request: Request,
    db: Session = Depends(get_db),
    page: int = 1,
    categoria: Optional[str] = None
):
    """Listar todos los contactos del usuario"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=303)
    
    print(f"\n=== LISTANDO CONTACTOS PARA USUARIO ID: {usuario_id} ===")
    print(f"Filtro recibido - categoria: '{categoria}'")
    
    data = crud.obtener_contactos_paginados(
        db,
        usuario_id,
        page=page,
        categoria=categoria
    )
    
    mensaje = request.session.pop("mensaje", None)
    
    return templates.TemplateResponse(
        "contactos_listado.html",
        {
            "request": request,
            "contactos": data["contactos"],
            "total_pages": data["total_pages"],
            "current_page": page,
            "filtro_categoria": categoria,
            "mensaje": mensaje
        }
    )



@router.get("/contactos/nuevo", response_class=HTMLResponse)
@router.get("/contactos/editar/{contacto_id}", response_class=HTMLResponse)
def formulario_contacto(
    request: Request,
    contacto_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """Mostrar formulario para crear/editar contacto"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=303)

    contacto = None
    if contacto_id:
        contacto = crud.obtener_contacto(db, contacto_id)
        if not contacto or contacto.usuario_id != usuario_id:
            raise HTTPException(status_code=404, detail="Contacto no encontrado")
    
    return templates.TemplateResponse("contactos_form.html", {
        "request": request,
        "contacto": contacto
    })

@router.post("/contactos/guardar")
def guardar_contacto(
    request: Request,
    id: Optional[int] = Form(None),
    nombres: str = Form(...),
    apellidos: str = Form(...),
    categoria: str = Form("otro"),
    direccion: Optional[str] = Form(None),
    celular1: str = Form(...),  # Ahora es obligatorio
    celular2: Optional[str] = Form(None),
    email: Optional[str] = Form(None),
    notas: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):

    """Guardar contacto (crear o actualizar)"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=303)

    try:
        print(f"\n{'='*60}")
        print(f"{'📝 EDITANDO CONTACTO' if id else '➕ CREANDO NUEVO CONTACTO'}")
        print(f"{'='*60}")
        
        # Validar datos obligatorios
        if not nombres.strip() or not apellidos.strip():
            raise ValueError("❌ ERROR: Nombres y apellidos son requeridos")
        
        # Validar categoría
        categorias_validas = ['familia', 'amigos', 'trabajo', 'servicios', 'educacion', 'otro']
        if categoria not in categorias_validas:
            categoria = 'otro'
        
        # Validar email si se proporciona
        email_clean = email.strip() if email else None
        if email_clean and "@" not in email_clean:
            email_clean = None
        
        # Preparar datos del contacto
        contacto_data = {
            "nombres": nombres.strip(),
            "apellidos": apellidos.strip(),
            "categoria": categoria,
            "direccion": direccion.strip() if direccion else None,
            "celular1": celular1.strip() if celular1 else None,
            "celular2": celular2.strip() if celular2 else None,
            "email": email_clean,
            "notas": notas.strip() if notas else None
        }
        
        if id:
            # Actualizar contacto existente
            contacto_update = schemas.ContactoUpdate(**contacto_data)
            crud.actualizar_contacto(db, contacto_id=int(id), contacto=contacto_update, usuario_id=usuario_id)
            mensaje = "Contacto actualizado correctamente"
        else:
            # Crear nuevo contacto
            contacto_create = schemas.ContactoCreate(**contacto_data)
            crud.crear_contacto(db, contacto=contacto_create, usuario_id=usuario_id)
            mensaje = "Contacto creado correctamente"

        request.session['mensaje'] = {
            'tipo': 'exito',
            'titulo': '¡Éxito!',
            'texto': mensaje
        }
        
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        db.rollback()
        request.session['mensaje'] = {
            'tipo': 'error',
            'titulo': 'Error',
            'texto': f'Ocurrió un error: {str(e)}'
        }

    return RedirectResponse(url="/contactos", status_code=303)

@router.get("/contactos/eliminar/{contacto_id}")
def eliminar_contacto(
    request: Request,
    contacto_id: int,
    db: Session = Depends(get_db)
):
    """Eliminar un contacto"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=303)

    resultado = crud.eliminar_contacto(db, contacto_id, usuario_id)
    
    if resultado:
        request.session["mensaje"] = {
            "tipo": "exito",
            "titulo": "¡Eliminado!",
            "texto": "Contacto eliminado correctamente"
        }
    else:
        request.session["mensaje"] = {
            "tipo": "error",
            "titulo": "Error",
            "texto": "No se pudo eliminar el contacto"
        }
    
    return RedirectResponse(url="/contactos", status_code=303)




 # Ingresos Descargar Excel

@router.get("/ingresos/descargar-excel")
def descargar_ingresos_excel(
    request: Request,
    db: Session = Depends(get_db),
    tipo: Optional[str] = None,
    estado: Optional[str] = None
):
    """Descarga todos los ingresos en formato Excel"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=303)
    
    print(f"\n=== DESCARGANDO INGRESOS A EXCEL ===")
    print(f"Usuario ID: {usuario_id}")
    print(f"Filtros - tipo: '{tipo}', estado: '{estado}'")
    
    # Obtener TODOS los ingresos (sin paginación)
    query = db.query(models.Ingreso).join(
        models.Categoria, models.Ingreso.categoria_id == models.Categoria.id
    ).filter(models.Ingreso.usuario_id == usuario_id)
    
    if tipo and tipo.strip():
        query = query.filter(models.Categoria.tipo == tipo.strip())
    
    if estado in ['recibido', 'pendiente']:
        query = query.filter(models.Ingreso.estado == estado)
    
    ingresos = query.order_by(models.Ingreso.fecha.desc()).all()
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ingresos"
    
    # Estilos
    header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Nombre", "Valor", "Estado", "Fecha", "Categoría", "Tipo"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # Datos
    for row_num, ingreso in enumerate(ingresos, 2):
        # Nombre (Notas)
        cell = ws.cell(row=row_num, column=1)
        cell.value = ingreso.notas or '-'
        cell.border = border
        
        # Valor
        cell = ws.cell(row=row_num, column=2)
        cell.value = float(ingreso.valor)
        cell.number_format = '$#,##0'
        cell.border = border
        
        # Estado
        cell = ws.cell(row=row_num, column=3)
        cell.value = "Recibido" if ingreso.estado == 'recibido' else "Pendiente"
        if ingreso.estado == 'recibido':
            cell.fill = PatternFill(start_color="d1fae5", end_color="d1fae5", fill_type="solid")
            cell.font = Font(color="065f46")
        else:
            cell.fill = PatternFill(start_color="fed7aa", end_color="fed7aa", fill_type="solid")
            cell.font = Font(color="92400e")
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
        
        # Fecha
        cell = ws.cell(row=row_num, column=4)
        cell.value = ingreso.fecha.strftime('%d/%m/%Y') if ingreso.fecha else '-'
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
        
        # Categoría
        cell = ws.cell(row=row_num, column=5)
        cell.value = ingreso.categoria.nombre
        cell.border = border
        
        # Tipo
        cell = ws.cell(row=row_num, column=6)
        cell.value = ingreso.categoria.tipo.capitalize()
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 12
    
    # Generar archivo en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Nombre del archivo
    from datetime import datetime
    fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"ingresos_{fecha_actual}.xlsx"
    
    print(f"✅ Excel generado: {len(ingresos)} registros")
    
    # Retornar archivo
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# Gastos Descargar Excel
@router.get("/gastos/descargar-excel")
def descargar_gastos_excel(
    request: Request,
    db: Session = Depends(get_db),
    tipo: Optional[str] = None,
    estado: Optional[str] = None
):
    """Descarga todos los gastos en formato Excel"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=303)

    print(f"\n=== DESCARGANDO GASTOS A EXCEL ===")
    print(f"Usuario ID: {usuario_id}")
    print(f"Filtros - tipo: '{tipo}', estado: '{estado}'")

    # Query base
    query = db.query(models.Gasto).join(
        models.Categoria, models.Gasto.categoria_id == models.Categoria.id
    ).filter(models.Gasto.usuario_id == usuario_id)

    if tipo and tipo.strip():
        query = query.filter(models.Categoria.tipo == tipo.strip())

    if estado in ['pagado', 'pendiente']:
        if estado == 'pagado':
            query = query.filter(models.Gasto.pagado == True)
        else:
            query = query.filter(models.Gasto.pagado == False)

    gastos = query.order_by(models.Gasto.fecha_limite.desc()).all()

    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gastos"

    # Estilos
    header_fill = PatternFill(start_color="dc2626", end_color="dc2626", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ["Nombre", "Valor", "Estado", "Fecha límite", "Categoría", "Tipo"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Datos
    for row_num, gasto in enumerate(gastos, 2):
        # Nombre
        cell = ws.cell(row=row_num, column=1)
        cell.value = gasto.notas or '-'
        cell.border = border

        # Valor
        cell = ws.cell(row=row_num, column=2)
        cell.value = float(gasto.valor)
        cell.number_format = '$#,##0'
        cell.border = border

        # Estado
        cell = ws.cell(row=row_num, column=3)
        cell.value = "Pagado" if gasto.pagado else "Pendiente"
        if gasto.pagado:
            cell.fill = PatternFill(start_color="d1fae5", end_color="d1fae5", fill_type="solid")
            cell.font = Font(color="065f46")
        else:
            cell.fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
            cell.font = Font(color="991b1b")
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

        # Fecha límite
        cell = ws.cell(row=row_num, column=4)
        cell.value = gasto.fecha_limite.strftime('%d/%m/%Y') if gasto.fecha_limite else '-'
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

        # Categoría
        cell = ws.cell(row=row_num, column=5)
        cell.value = gasto.categoria.nombre
        cell.border = border

        # Tipo
        cell = ws.cell(row=row_num, column=6)
        cell.value = gasto.categoria.tipo.capitalize()
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    # Ajustar columnas
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 12

    # Generar archivo
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    from datetime import datetime
    fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"gastos_{fecha_actual}.xlsx"

    print(f"✅ Excel generado: {len(gastos)} registros")

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )




# ============================================================================
# 💰 RUTAS DE PAGOS - AGREGAR EN routes.py
# ============================================================================

@router.get("/pagos/nuevo/{credito_id}", response_class=HTMLResponse)
def formulario_nuevo_pago(
    request: Request,
    credito_id: int,
    db: Session = Depends(get_db)
):
    """Mostrar formulario para crear nuevo pago"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=303)
    
    # Verificar que el crédito existe y pertenece al usuario
    credito = crud.obtener_credito(db, credito_id)
    if not credito or credito.usuario_id != usuario_id:
        raise HTTPException(status_code=404, detail="Crédito no encontrado")
    
    # ✅ CALCULAR TOTAL PAGADO PARA MOSTRAR EN EL RESUMEN
    pagos = crud.obtener_pagos_por_credito(db, credito_id)
    total_pagado = sum(float(pago.monto) for pago in pagos) if pagos else 0
    
    # Fecha actual para el formulario
    hoy = date.today().isoformat()
    
    return templates.TemplateResponse("pago_form.html", {
        "request": request,
        "credito": credito,
        "total_pagado": total_pagado,  # ✅ ESTO FALTABA
        "hoy": hoy
    })



@router.post("/pagos/guardar")
def guardar_pago(
    request: Request,
    credito_id: int = Form(...),
    monto: str = Form(...),
    fecha_pago: str = Form(...),
    comprobante: str = Form(...),
    notas: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    """Guardar un nuevo pago"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        print(f"\n{'='*60}")
        print(f"💰 PROCESANDO PAGO")
        print(f"{'='*60}")
        print(f"  Crédito ID: {credito_id}")
        print(f"  Monto recibido (string): '{monto}'")
        print(f"  Fecha: {fecha_pago}")
        print(f"  Comprobante: {comprobante}")
        
        # ✅ 1. LIMPIAR EL MONTO - ELIMINAR TODO LO QUE NO SEA NÚMERO
        monto_limpio = ''.join(c for c in monto if c.isdigit())
        
        if not monto_limpio:
            raise ValueError("El monto no contiene números válidos")
        
        # ✅ 2. CONVERTIR A FLOAT (SIN DIVIDIR ENTRE 100)
        monto_float = float(monto_limpio)
        print(f"  Monto limpiado: '{monto_limpio}' -> {monto_float:,.2f}")
        
        # ✅ 3. VALIDAR QUE EL CRÉDITO EXISTE
        credito = crud.obtener_credito(db, credito_id)
        if not credito or credito.usuario_id != usuario_id:
            raise HTTPException(status_code=404, detail="Crédito no encontrado")
        
        print(f"  Saldo actual: ${credito.saldo_actual:,.2f}")
        
        # ✅ 4. VALIDAR QUE EL MONTO NO EXCEDA EL SALDO
        if monto_float > credito.saldo_actual:
            raise ValueError(
                f"El monto del pago (${monto_float:,.0f}) "
                f"excede el saldo actual (${credito.saldo_actual:,.0f})"
            )
        
        # ✅ 5. CONVERTIR FECHA
        try:
            fecha_pago_dt = date.fromisoformat(fecha_pago)
        except ValueError:
            raise ValueError(f"Formato de fecha inválido: {fecha_pago}")
        
        # ✅ 6. CREAR EL PAGO
        pago_data = schemas.PagoCreate(
            credito_id=credito_id,
            monto=monto_float,
            fecha_pago=fecha_pago_dt,
            comprobante=comprobante.strip(),
            notas=notas.strip() if notas else None
        )
        
        pago = crud.crear_pago(db, pago_data, usuario_id)
        
        if pago:
            mensaje = f'✅ Pago de ${monto_float:,.0f} registrado correctamente'
            print(mensaje)
            request.session['mensaje'] = {
                'tipo': 'success',
                'titulo': '¡Éxito!',
                'texto': mensaje
            }
        else:
            raise Exception("No se pudo crear el pago en la base de datos")
        
    except ValueError as e:
        print(f"❌ Error de validación: {e}")
        request.session['mensaje'] = {
            'tipo': 'error',
            'titulo': 'Error de validación',
            'texto': str(e)
        }
    except HTTPException as e:
        print(f"❌ HTTP Exception: {e.detail}")
        request.session['mensaje'] = {
            'tipo': 'error',
            'titulo': 'Error',
            'texto': e.detail
        }
    except Exception as e:
        print(f"❌ Error inesperado: {e}")
        import traceback
        traceback.print_exc()
        db.rollback()
        request.session['mensaje'] = {
            'tipo': 'error',
            'titulo': 'Error del servidor',
            'texto': f'Error al registrar pago: {str(e)[:100]}'
        }
    
    return RedirectResponse(url=f"/creditos/detalle/{credito_id}", status_code=303)



@router.get("/pagos/eliminar/{pago_id}")
def eliminar_pago_route(
    request: Request,
    pago_id: int,
    credito_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """Eliminar un pago"""
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return RedirectResponse(url="/login", status_code=303)
    
    resultado = crud.eliminar_pago(db, pago_id, usuario_id)
    
    if resultado:
        request.session["mensaje"] = {
            "tipo": "success",  # ✅ Cambiado de 'exito' a 'success' para SweetAlert
            "titulo": "¡Eliminado!",
            "texto": "Pago eliminado correctamente"
        }
        print(f"✅ Pago {pago_id} eliminado correctamente")
    else:
        request.session["mensaje"] = {
            "tipo": "error",
            "titulo": "Error",
            "texto": "No se pudo eliminar el pago"
        }
        print(f"❌ Error al eliminar pago {pago_id}")
    
    # Redirigir al detalle del crédito
    if credito_id:
        return RedirectResponse(url=f"/creditos/detalle/{credito_id}", status_code=303)
    else:
        # Si no viene credito_id, obtenerlo del pago
        pago = db.query(models.Pago).filter(models.Pago.id == pago_id).first()
        if pago:
            return RedirectResponse(url=f"/creditos/detalle/{pago.credito_id}", status_code=303)
        else:
            return RedirectResponse(url="/creditos", status_code=303)